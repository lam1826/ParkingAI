from __future__ import annotations

from datetime import date
from html import escape
from io import BytesIO
from pathlib import Path
import re
from typing import Any, Dict, Iterable, Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

from core.clock import business_today
from services.report_service import ReportService


_XML_ILLEGAL_CHARACTERS = re.compile(
    r"[\x00-\x08\x0B\x0C\x0E-\x1F\uD800-\uDFFF\uFFFE\uFFFF]"
)
_MAX_EXACT_EXCEL_INTEGER = 999_999_999_999_999


class ReportExportService:
    """Tạo tệp báo cáo từ cùng nguồn dữ liệu đang hiển thị trên giao diện."""

    def __init__(self, db: Session):
        self.report_service = ReportService(db)

    def _get_data(
        self,
        period: Literal["day", "week", "month", "year"],
        anchor_date: date | None = None,
    ) -> Dict[str, Any]:
        resolved_anchor = anchor_date or business_today()
        return {
            "summary": self.report_service.get_revenue_report(
                period, anchor_date=resolved_anchor
            ),
            "traffic": self.report_service.get_traffic_report(
                period, anchor_date=resolved_anchor
            ),
        }

    @staticmethod
    def _excel_safe_text(value: Any) -> Any:
        """Return XML-safe text that cannot become a spreadsheet formula."""
        if isinstance(value, str):
            value = _XML_ILLEGAL_CHARACTERS.sub("", value)
            if value.startswith(("=", "+", "-", "@", "\t", "\r")):
                return "'" + value
        return value

    @classmethod
    def _excel_safe_value(cls, value: Any, *, monetary: bool = False) -> Any:
        """Keep large money exact despite Excel's 15-digit numeric limit."""
        if (
            monetary
            and isinstance(value, (int, float))
            and not isinstance(value, bool)
            and abs(value) > _MAX_EXACT_EXCEL_INTEGER
        ):
            value = str(value)
        return cls._excel_safe_text(value)

    @staticmethod
    def _style_header(row: Iterable) -> None:
        fill = PatternFill("solid", fgColor="1976D2")
        for cell in row:
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

    def build_excel(
        self,
        period: Literal["day", "week", "month", "year"],
        anchor_date: date | None = None,
    ) -> bytes:
        data = self._get_data(period, anchor_date=anchor_date)
        summary = data["summary"]
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Tong quan"

        sheet.append(["BÁO CÁO BÃI ĐỖ XE"])
        sheet.merge_cells("A1:B1")
        sheet["A1"].font = Font(size=16, bold=True, color="1976D2")
        sheet["A1"].alignment = Alignment(horizontal="center")
        sheet.append(["Chỉ số", "Giá trị"])
        self._style_header(sheet[2])
        rows = [
            ("Kỳ báo cáo", summary["filter_type"], False),
            ("Từ ngày", summary["start_date"].strftime("%d/%m/%Y %H:%M"), False),
            ("Đến ngày", summary["end_date"].strftime("%d/%m/%Y %H:%M"), False),
            ("Tổng lượt xe", summary["total_trips"], False),
            ("Tổng doanh thu", summary["total_revenue"], True),
            ("Phí trung bình", summary["average_fee"], True),
            ("Loại xe phổ biến", summary["most_frequent_vehicle_type"], False),
        ]
        for label, value, monetary in rows:
            sheet.append((
                self._excel_safe_text(label),
                self._excel_safe_value(value, monetary=monetary),
            ))
        sheet["B7"].number_format = '#,##0 "₫"'
        sheet["B8"].number_format = '#,##0 "₫"'
        sheet.column_dimensions["A"].width = 24
        sheet.column_dimensions["B"].width = 32

        traffic_sheets = [
            ("Theo gio", "traffic_by_hour"),
            ("Theo ngay", "traffic_by_day"),
            ("Theo tuan", "traffic_by_week"),
            ("Theo thang", "traffic_by_month"),
        ]
        for title, key in traffic_sheets:
            traffic_sheet = workbook.create_sheet(title)
            traffic_sheet.append(["Thời gian", "Tổng lượt xe"])
            self._style_header(traffic_sheet[1])
            for item in data["traffic"][key]:
                traffic_sheet.append([
                    self._excel_safe_text(item["time_label"]),
                    item["total_vehicles"],
                ])
            traffic_sheet.column_dimensions["A"].width = 24
            traffic_sheet.column_dimensions["B"].width = 18
            traffic_sheet.freeze_panes = "A2"

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    @staticmethod
    def _register_unicode_font() -> str:
        font_candidates = [
            Path("C:/Windows/Fonts/arial.ttf"),
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        ]
        for path in font_candidates:
            if path.exists():
                font_name = "ParkingAIUnicode"
                if font_name not in pdfmetrics.getRegisteredFontNames():
                    pdfmetrics.registerFont(TTFont(font_name, str(path)))
                return font_name
        return "Helvetica"

    def build_pdf(
        self,
        period: Literal["day", "week", "month", "year"],
        anchor_date: date | None = None,
    ) -> bytes:
        data = self._get_data(period, anchor_date=anchor_date)
        summary = data["summary"]
        output = BytesIO()
        font_name = self._register_unicode_font()
        document = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=16 * mm,
            leftMargin=16 * mm,
            topMargin=14 * mm,
            bottomMargin=14 * mm,
            title="Báo cáo bãi đỗ xe ParkingAI",
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "ParkingTitle",
            parent=styles["Title"],
            fontName=font_name,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#1976D2"),
        )
        body_style = ParagraphStyle("ParkingBody", parent=styles["BodyText"], fontName=font_name, leading=15)
        heading_style = ParagraphStyle("ParkingHeading", parent=styles["Heading2"], fontName=font_name, textColor=colors.HexColor("#1976D2"))
        story = [Paragraph("BÁO CÁO BÃI ĐỖ XE PARKINGAI", title_style), Spacer(1, 6 * mm)]

        summary_rows = [
            ["Chỉ số", "Giá trị"],
            ["Kỳ báo cáo", summary["filter_type"]],
            ["Khoảng thời gian", f'{summary["start_date"].strftime("%d/%m/%Y")} - {summary["end_date"].strftime("%d/%m/%Y")}'],
            ["Tổng lượt xe", str(summary["total_trips"])],
            ["Tổng doanh thu", f'{summary["total_revenue"]:,.0f} ₫'],
            ["Phí trung bình", f'{summary["average_fee"]:,.0f} ₫'],
            ["Loại xe phổ biến", escape(str(summary["most_frequent_vehicle_type"]))],
        ]
        summary_table = Table(summary_rows, colWidths=[55 * mm, 105 * mm], repeatRows=1)
        summary_table.setStyle(self._pdf_table_style(font_name))
        story.extend([summary_table, Spacer(1, 7 * mm)])

        sections = [
            ("Lưu lượng theo giờ", "traffic_by_hour"),
            ("Lưu lượng theo ngày", "traffic_by_day"),
            ("Lưu lượng theo tuần", "traffic_by_week"),
            ("Lưu lượng theo tháng", "traffic_by_month"),
        ]
        for title, key in sections:
            story.append(Paragraph(title, heading_style))
            items = data["traffic"][key]
            if not items:
                story.append(Paragraph("Chưa có dữ liệu.", body_style))
            else:
                table_rows = [["Thời gian", "Tổng lượt xe"]] + [
                    [escape(str(item["time_label"])), str(item["total_vehicles"])] for item in items
                ]
                table = Table(table_rows, colWidths=[100 * mm, 60 * mm], repeatRows=1)
                table.setStyle(self._pdf_table_style(font_name))
                story.append(table)
            story.append(Spacer(1, 5 * mm))

        document.build(story)
        return output.getvalue()

    @staticmethod
    def _pdf_table_style(font_name: str) -> TableStyle:
        return TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1976D2")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B0BEC5")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ])
