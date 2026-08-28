import datetime as dt
from datetime import date, datetime, timedelta, timezone
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.orm import Session

import core.clock as clock_module
from core.money import ExactVndRangeError, MAX_EXACT_VND
from models.parking_session import ParkingSession
from services.auth_service import AuthService
from services.parking_service import ParkingService
from services.report_export_service import ReportExportService


class FixedBusinessClock(dt.datetime):
    """Một instant cố định: 27/08/2026 12:00 giờ Việt Nam."""

    FIXED_UTC = datetime(2026, 8, 27, 5, 0, tzinfo=timezone.utc)

    @classmethod
    def now(cls, tz=None):
        if tz is None:
            return cls.FIXED_UTC.replace(tzinfo=None)
        return cls.FIXED_UTC.astimezone(tz)


def _headers(user) -> dict[str, str]:
    token = AuthService().create_access_token(
        user_id=user.id,
        username=user.username,
        role=user.role.name,
    )
    return {"Authorization": f"Bearer {token}"}


def _completed_session(
    db: Session,
    *,
    vehicle_id: int,
    staff_id: int,
    check_in_time: datetime,
    parking_fee: int = 10_000,
) -> ParkingSession:
    session = ParkingSession(
        vehicle_id=vehicle_id,
        parking_slot_id=None,
        check_in_time=check_in_time,
        check_out_time=check_in_time + timedelta(hours=1),
        parking_fee=parking_fee,
        status="completed",
        staff_in_id=staff_id,
        staff_out_id=staff_id,
    )
    db.add(session)
    return session


def test_public_revenue_aggregates_fail_closed_above_exact_vnd_range(
    monkeypatch,
    client: TestClient,
    db_session: Session,
    test_user,
    vehicle,
):
    """A JSON number above 2**53-1 would be rounded by every JS client.

    The two persisted fees are individually valid, but their known total is
    MAX_EXACT_VND + 2. Every public aggregate must reject that total instead
    of publishing a subtly different amount to the browser/export pipeline.
    """
    monkeypatch.setattr(clock_module, "datetime", FixedBusinessClock)
    _completed_session(
        db_session,
        vehicle_id=vehicle.id,
        staff_id=test_user.id,
        check_in_time=datetime(2026, 8, 27, 8, 15),
        parking_fee=MAX_EXACT_VND,
    )
    _completed_session(
        db_session,
        vehicle_id=vehicle.id,
        staff_id=test_user.id,
        check_in_time=datetime(2026, 8, 27, 10, 15),
        parking_fee=2,
    )
    db_session.commit()

    headers = _headers(test_user)
    for endpoint in (
        "/reports/revenue?period=day",
        "/parking/statistics",
        "/dashboard",
        "/dashboard/revenue-chart",
    ):
        response = client.get(endpoint, headers=headers)
        assert response.status_code == 500, (endpoint, response.text)
        assert response.json()["detail"] == (
            "Tổng doanh thu vượt phạm vi VND chính xác được hỗ trợ."
        )


def test_daily_summaries_fail_closed_above_exact_vnd_range(
    db_session: Session,
    test_user,
    vehicle,
):
    _completed_session(
        db_session,
        vehicle_id=vehicle.id,
        staff_id=test_user.id,
        check_in_time=datetime(2026, 8, 27, 8, 15),
        parking_fee=MAX_EXACT_VND,
    )
    _completed_session(
        db_session,
        vehicle_id=vehicle.id,
        staff_id=test_user.id,
        check_in_time=datetime(2026, 8, 27, 10, 15),
        parking_fee=2,
    )
    db_session.commit()

    with pytest.raises(
        ExactVndRangeError,
        match="Tổng doanh thu vượt phạm vi VND chính xác được hỗ trợ",
    ):
        ParkingService(db_session).get_daily_summaries(
            date(2026, 8, 27),
            date(2026, 8, 27),
        )


def test_revenue_aggregates_over_sqlite_int64_fail_with_domain_error(
    monkeypatch,
    client: TestClient,
    db_session: Session,
    test_user,
    vehicle,
):
    """A valid row set must never leak SQLite's ``integer overflow``.

    1024 * (2**53 - 1) still fits signed int64; the 1025th row crosses that
    database limit.  Every public aggregation path must nevertheless reach
    the exact-VND domain gate and return its stable error contract.
    """
    monkeypatch.setattr(clock_module, "datetime", FixedBusinessClock)
    for index in range(1025):
        _completed_session(
            db_session,
            vehicle_id=vehicle.id,
            staff_id=test_user.id,
            check_in_time=datetime(2026, 8, 27, 8, 0) + timedelta(seconds=index),
            parking_fee=MAX_EXACT_VND,
        )
    db_session.commit()

    expected_detail = "Tổng doanh thu vượt phạm vi VND chính xác được hỗ trợ."
    headers = _headers(test_user)
    for endpoint in (
        "/reports/revenue?period=day",
        "/parking/statistics",
        "/dashboard",
        "/dashboard/revenue-chart",
    ):
        response = client.get(endpoint, headers=headers)
        assert response.status_code == 500, (endpoint, response.text)
        assert response.json() == {"detail": expected_detail}
        assert "integer overflow" not in response.text.lower()

    with pytest.raises(ExactVndRangeError, match=expected_detail[:-1]):
        ParkingService(db_session).get_daily_summaries(
            date(2026, 8, 27),
            date(2026, 8, 27),
        )


def test_revenue_last_7_days_excludes_future_rows_at_upper_boundary(
    monkeypatch,
    db_session: Session,
    test_user,
    vehicle,
):
    """Future completed rows cannot poison the current seven-day window."""
    monkeypatch.setattr(clock_module, "datetime", FixedBusinessClock)
    _completed_session(
        db_session,
        vehicle_id=vehicle.id,
        staff_id=test_user.id,
        check_in_time=datetime(2026, 8, 27, 8, 0),
        parking_fee=12_345,
    )
    # First future checkout is exactly the half-open upper bound (28/08
    # 00:00); together the future group exceeds MAX_EXACT_VND. Without the
    # end-exclusive filter it raises and breaks an otherwise valid chart.
    _completed_session(
        db_session,
        vehicle_id=vehicle.id,
        staff_id=test_user.id,
        check_in_time=datetime(2026, 8, 27, 23, 0),
        parking_fee=MAX_EXACT_VND,
    )
    _completed_session(
        db_session,
        vehicle_id=vehicle.id,
        staff_id=test_user.id,
        check_in_time=datetime(2026, 8, 27, 23, 10),
        parking_fee=1,
    )
    db_session.commit()

    result = ParkingService(db_session).get_revenue_last_7_days()

    assert len(result) == 7
    assert result[-1] == {"day": "27/08", "revenue": 12_345}
    assert sum(item["revenue"] for item in result) == 12_345


@pytest.mark.parametrize(
    ("period", "inside", "outside"),
    [
        ("day", datetime(2026, 8, 27, 8, 15), datetime(2026, 8, 26, 8, 15)),
        ("week", datetime(2026, 8, 24, 8, 15), datetime(2026, 8, 23, 8, 15)),
        ("month", datetime(2026, 8, 1, 8, 15), datetime(2026, 7, 31, 8, 15)),
        ("year", datetime(2026, 1, 1, 8, 15), datetime(2025, 12, 31, 8, 15)),
    ],
)
def test_traffic_endpoint_filters_every_aggregation_to_selected_period(
    monkeypatch,
    client: TestClient,
    db_session: Session,
    test_user,
    vehicle,
    period: str,
    inside: datetime,
    outside: datetime,
):
    """Traffic/giờ cao điểm không được cộng bản ghi all-time ngoài kỳ.

    Hai session cố ý cùng giờ 08:15: nếu query còn all-time thì bucket 08:00
    sẽ có count=2 và test chắc chắn đỏ, không thể pass do khác nhãn thời gian.
    """
    monkeypatch.setattr(clock_module, "datetime", FixedBusinessClock)
    _completed_session(
        db_session,
        vehicle_id=vehicle.id,
        staff_id=test_user.id,
        check_in_time=inside,
    )
    _completed_session(
        db_session,
        vehicle_id=vehicle.id,
        staff_id=test_user.id,
        check_in_time=outside,
    )
    db_session.commit()

    response = client.get(
        "/reports/traffic",
        params={"period": period},
        headers=_headers(test_user),
    )

    assert response.status_code == 200
    traffic = response.json()
    assert traffic["traffic_by_hour"] == [
        {"time_label": "08:00", "total_vehicles": 1}
    ]
    assert traffic["traffic_by_day"] == [
        {"time_label": inside.strftime("%Y-%m-%d"), "total_vehicles": 1}
    ]
    assert sum(item["total_vehicles"] for item in traffic["traffic_by_week"]) == 1
    assert sum(item["total_vehicles"] for item in traffic["traffic_by_month"]) == 1


def test_excel_export_traffic_uses_same_selected_period_as_revenue(
    monkeypatch,
    client: TestClient,
    db_session: Session,
    test_user,
    vehicle,
):
    monkeypatch.setattr(clock_module, "datetime", FixedBusinessClock)
    _completed_session(
        db_session,
        vehicle_id=vehicle.id,
        staff_id=test_user.id,
        check_in_time=datetime(2026, 8, 27, 8, 15),
    )
    _completed_session(
        db_session,
        vehicle_id=vehicle.id,
        staff_id=test_user.id,
        check_in_time=datetime(2026, 8, 26, 8, 15),
    )
    db_session.commit()

    response = client.get(
        "/reports/export/xlsx",
        params={"period": "day"},
        headers=_headers(test_user),
    )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    hourly_rows = list(workbook["Theo gio"].iter_rows(values_only=True))
    assert hourly_rows == [
        ("Thời gian", "Tổng lượt xe"),
        ("08:00", 1),
    ]


def test_report_endpoints_honor_one_explicit_anchor_date(
    client: TestClient,
    db_session: Session,
    test_user,
    vehicle,
):
    """Two HTTP calls can cross midnight; anchor_date keeps their period identical."""
    _completed_session(
        db_session,
        vehicle_id=vehicle.id,
        staff_id=test_user.id,
        check_in_time=datetime(2026, 9, 1, 8, 15),
    )
    db_session.commit()
    params = {"period": "day", "anchor_date": "2026-09-01"}
    headers = _headers(test_user)

    revenue = client.get("/reports/revenue", params=params, headers=headers)
    traffic = client.get("/reports/traffic", params=params, headers=headers)

    assert revenue.status_code == 200
    assert traffic.status_code == 200
    assert revenue.json()["start_date"].startswith("2026-09-01T00:00:00")
    assert revenue.json()["total_trips"] == 1
    assert traffic.json()["traffic_by_hour"] == [
        {"time_label": "08:00", "total_vehicles": 1}
    ]


def test_export_honors_explicit_anchor_in_content_and_filename(
    client: TestClient,
    db_session: Session,
    test_user,
    vehicle,
):
    _completed_session(
        db_session,
        vehicle_id=vehicle.id,
        staff_id=test_user.id,
        check_in_time=datetime(2026, 8, 31, 8, 15),
    )
    _completed_session(
        db_session,
        vehicle_id=vehicle.id,
        staff_id=test_user.id,
        check_in_time=datetime(2026, 9, 1, 9, 15),
    )
    db_session.commit()

    response = client.get(
        "/reports/export/xlsx",
        params={"period": "day", "anchor_date": "2026-09-01"},
        headers=_headers(test_user),
    )

    assert response.status_code == 200
    assert "parking-report-day-2026-09-01.xlsx" in response.headers[
        "content-disposition"
    ]
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    assert list(workbook["Theo gio"].iter_rows(values_only=True)) == [
        ("Thời gian", "Tổng lượt xe"),
        ("09:00", 1),
    ]


@pytest.mark.parametrize(
    ("endpoint", "period"),
    [
        ("/reports/revenue", "day"),
        ("/reports/revenue", "week"),
        ("/reports/revenue", "month"),
        ("/reports/revenue", "year"),
        ("/reports/traffic", "day"),
        ("/reports/export/xlsx", "day"),
    ],
)
def test_report_endpoints_reject_anchor_dates_that_overflow_period_bounds(
    client: TestClient,
    test_user,
    endpoint: str,
    period: str,
):
    response = client.get(
        endpoint,
        params={"period": period, "anchor_date": "9999-12-31"},
        headers=_headers(test_user),
    )

    assert response.status_code == 422
    assert response.json()["detail"] == (
        "Ngày neo nằm ngoài phạm vi báo cáo được hỗ trợ."
    )


def test_export_samples_business_date_once_for_summary_and_traffic(
    db_session: Session,
    monkeypatch,
):
    service = ReportExportService(db_session)
    anchors = []

    monkeypatch.setattr(
        "services.report_export_service.business_today",
        lambda: date(2026, 8, 31),
        raising=False,
    )
    monkeypatch.setattr(
        service.report_service,
        "get_revenue_report",
        lambda period, anchor_date=None: anchors.append(anchor_date) or {},
    )
    monkeypatch.setattr(
        service.report_service,
        "get_traffic_report",
        lambda period, anchor_date=None: anchors.append(anchor_date) or {},
    )

    service._get_data("month")

    assert anchors == [date(2026, 8, 31), date(2026, 8, 31)]


def test_excel_export_neutralizes_formula_like_vehicle_type_name(
    monkeypatch,
    client: TestClient,
    db_session: Session,
    test_user,
    vehicle,
):
    monkeypatch.setattr(clock_module, "datetime", FixedBusinessClock)
    vehicle.vehicle_type.name = '=HYPERLINK("https://invalid.example","click")'
    _completed_session(
        db_session,
        vehicle_id=vehicle.id,
        staff_id=test_user.id,
        check_in_time=datetime(2026, 8, 27, 8, 15),
    )
    db_session.commit()

    response = client.get(
        "/reports/export/xlsx",
        params={"period": "day"},
        headers=_headers(test_user),
    )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), data_only=False)
    vehicle_type_cell = workbook["Tong quan"]["B9"]
    assert vehicle_type_cell.value.startswith("'=")
    assert vehicle_type_cell.data_type != "f"


def test_excel_export_removes_xml_illegal_vehicle_type_characters(
    monkeypatch,
    client: TestClient,
    db_session: Session,
    test_user,
    vehicle,
):
    monkeypatch.setattr(clock_module, "datetime", FixedBusinessClock)
    vehicle.vehicle_type.name = "Ô\x00tô\x0b điện"
    _completed_session(
        db_session,
        vehicle_id=vehicle.id,
        staff_id=test_user.id,
        check_in_time=datetime(2026, 8, 27, 8, 15),
    )
    db_session.commit()

    response = client.get(
        "/reports/export/xlsx",
        params={"period": "day", "anchor_date": "2026-08-27"},
        headers=_headers(test_user),
    )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), data_only=False)
    assert workbook["Tong quan"]["B9"].value == "Ôtô điện"


def test_excel_export_writes_large_exact_vnd_as_text(
    monkeypatch,
    client: TestClient,
    db_session: Session,
    test_user,
    vehicle,
):
    """Excel numeric cells only preserve 15 significant decimal digits."""
    monkeypatch.setattr(clock_module, "datetime", FixedBusinessClock)
    _completed_session(
        db_session,
        vehicle_id=vehicle.id,
        staff_id=test_user.id,
        check_in_time=datetime(2026, 8, 27, 8, 15),
        parking_fee=MAX_EXACT_VND,
    )
    db_session.commit()

    response = client.get(
        "/reports/export/xlsx",
        params={"period": "day", "anchor_date": "2026-08-27"},
        headers=_headers(test_user),
    )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), data_only=False)
    revenue_cell = workbook["Tong quan"]["B7"]
    assert revenue_cell.value == str(MAX_EXACT_VND)
    assert revenue_cell.data_type == "s"


def test_revenue_totals_keep_exact_integer_vnd_across_report_and_dashboard(
    monkeypatch,
    client: TestClient,
    db_session: Session,
    test_user,
    vehicle,
):
    """Giá trị đúng tại biên 2**53-1 vẫn phải được công bố chính xác."""
    monkeypatch.setattr(clock_module, "datetime", FixedBusinessClock)
    first = _completed_session(
        db_session,
        vehicle_id=vehicle.id,
        staff_id=test_user.id,
        check_in_time=datetime(2026, 8, 27, 8, 15),
    )
    second = _completed_session(
        db_session,
        vehicle_id=vehicle.id,
        staff_id=test_user.id,
        check_in_time=datetime(2026, 8, 27, 10, 15),
    )
    first.parking_fee = MAX_EXACT_VND - 1
    second.parking_fee = 1
    db_session.commit()

    expected = MAX_EXACT_VND
    report = client.get(
        "/reports/revenue",
        params={"period": "day"},
        headers=_headers(test_user),
    )
    dashboard = client.get("/dashboard", headers=_headers(test_user))

    assert report.status_code == 200
    assert dashboard.status_code == 200
    assert report.json()["total_revenue"] == expected
    assert type(report.json()["total_revenue"]) is int
    assert dashboard.json()["total_revenue_today"] == expected
    assert type(dashboard.json()["total_revenue_today"]) is int

    schemas = client.get("/openapi.json").json()["components"]["schemas"]
    assert schemas["RevenueReportResponse"]["properties"]["total_revenue"][
        "type"
    ] == "integer"
    assert schemas["DashboardResponse"]["properties"]["total_revenue_today"][
        "type"
    ] == "integer"
