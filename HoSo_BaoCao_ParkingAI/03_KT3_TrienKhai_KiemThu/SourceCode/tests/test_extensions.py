from io import BytesIO

import bcrypt
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from models.audit_log import AuditLog
from models.role import Role
from models.user import User
from services.auth_service import AuthService


def make_headers(user: User) -> dict[str, str]:
    token = AuthService().create_access_token(
        user_id=user.id,
        username=user.username,
        role=user.role.name,
    )
    return {"Authorization": f"Bearer {token}"}


def create_manager(db_session: Session) -> User:
    role = Role(name="manager", description="Quản lý")
    db_session.add(role)
    db_session.flush()
    manager = User(
        username="extension_manager",
        password_hash=bcrypt.hashpw(b"password123", bcrypt.gensalt()).decode("utf-8"),
        full_name="Quản lý mở rộng",
        role_id=role.id,
        is_active=True,
    )
    db_session.add(manager)
    db_session.commit()
    db_session.refresh(manager)
    return manager


def test_export_report_as_real_excel(client: TestClient, test_user: User):
    response = client.get(
        "/reports/export/xlsx",
        params={"period": "week"},
        headers=make_headers(test_user),
    )

    assert response.status_code == 200
    assert response.content.startswith(b"PK")
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert ".xlsx" in response.headers["content-disposition"]
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    assert workbook.sheetnames == ["Tong quan", "Theo gio", "Theo ngay", "Theo tuan", "Theo thang"]
    assert workbook["Tong quan"]["A1"].value == "BÁO CÁO BÃI ĐỖ XE"


def test_export_report_as_real_pdf(client: TestClient, test_user: User):
    response = client.get(
        "/reports/export/pdf",
        params={"period": "month"},
        headers=make_headers(test_user),
    )

    assert response.status_code == 200
    assert response.content.startswith(b"%PDF")
    assert response.headers["content-type"].startswith("application/pdf")
    assert ".pdf" in response.headers["content-disposition"]


def test_audit_middleware_records_authenticated_mutation(
    client: TestClient,
    db_session: Session,
    test_user: User,
):
    response = client.post(
        "/api/v1/zones",
        headers=make_headers(test_user),
        json={"name": "Khu nhật ký", "capacity": 20, "is_active": True},
    )

    assert response.status_code == 201
    audit = db_session.query(AuditLog).filter(AuditLog.path == "/api/v1/zones").one()
    assert audit.username == test_user.username
    assert audit.action == "CREATE"
    assert audit.resource == "zones"
    assert audit.success is True
    assert audit.status_code == 201


def test_only_manager_can_read_audit_logs(
    client: TestClient,
    db_session: Session,
    test_user: User,
):
    forbidden = client.get("/api/v1/audit-logs", headers=make_headers(test_user))
    assert forbidden.status_code == 403

    manager = create_manager(db_session)
    allowed = client.get("/api/v1/audit-logs", headers=make_headers(manager))
    assert allowed.status_code == 200
    assert isinstance(allowed.json(), list)
